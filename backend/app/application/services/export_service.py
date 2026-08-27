from __future__ import annotations
import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Tuple

from app.core.constants import EXPORT_MAX_ROWS
from app.core.exceptions import BadRequestError
from app.shared.enums import ReportExportFormat
from app.shared.utils.helpers import safe_round

_CSV_CONTENT_TYPE = "text/csv"
_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_CONTENT_TYPE = "application/pdf"


def _to_export_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime,)):
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, (int, float)):
        return value
    return str(value)


class ExportService:
    """Serializes tabular data into CSV, XLSX or PDF byte streams."""

    def export(
        self,
        *,
        report_type: str,
        rows: List[Dict[str, Any]],
        format: ReportExportFormat,
        title: str,
    ) -> Tuple[str, bytes, str]:
        if len(rows) > EXPORT_MAX_ROWS:
            raise BadRequestError(f"Too many rows to export (max {EXPORT_MAX_ROWS}).")
        if not rows:
            columns: List[str] = []
        else:
            columns = list(rows[0].keys())

        handler = {
            ReportExportFormat.CSV: self._to_csv,
            ReportExportFormat.XLSX: self._to_xlsx,
            ReportExportFormat.PDF: self._to_pdf,
        }.get(format)
        if handler is None:
            raise BadRequestError("Unsupported export format.")

        content = handler(rows=rows, columns=columns, title=title)
        content_type = {
            ReportExportFormat.CSV: _CSV_CONTENT_TYPE,
            ReportExportFormat.XLSX: _XLSX_CONTENT_TYPE,
            ReportExportFormat.PDF: _PDF_CONTENT_TYPE,
        }[format]
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_{stamp}.{format.value}"
        return content_type, content, filename

    def _to_csv(self, *, rows: List[Dict[str, Any]], columns: List[str], title: str) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        if columns:
            writer.writerow(columns)
        for row in rows:
            writer.writerow([_to_export_value(row.get(col)) for col in columns])
        return buffer.getvalue().encode("utf-8")

    def _to_xlsx(self, *, rows: List[Dict[str, Any]], columns: List[str], title: str) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = (title[:28] or "Report")

        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        if columns:
            for col_idx, col in enumerate(columns, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
                cell.fill = header_fill
                cell.font = header_font
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, col in enumerate(columns, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=_to_export_value(row.get(col)))

        for col_idx in range(1, len(columns) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 22

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _to_pdf(self, *, rows: List[Dict[str, Any]], columns: List[str], title: str) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()

        elements = []
        elements.append(Paragraph(title, styles["Title"]))
        elements.append(Spacer(1, 12))

        header = [col.replace("_", " ").title() for col in columns]
        table_data = [header]
        for row in rows:
            table_data.append([_to_export_value(row.get(col)) for col in columns])

        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)
        return buffer.getvalue()
